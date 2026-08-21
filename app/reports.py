from io import BytesIO
from pathlib import Path

from flask import Blueprint, abort, render_template, request, send_file
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select

from app.auth import roles_required
from app.extensions import db
from app.models import (
    Department,
    Plan,
    PlanItem,
    Publication,
    ReferenceValue,
)


REPORT_PUBLICATION_STATUSES = {
    "PREPARATION": "Подготовка статьи",
    "SUBMITTED_TO_JOURNAL": "Отправлена в журнал",
    "UNDER_REVIEW": "На рецензировании",
    "REVISION_REQUIRED": "Требуется доработка",
    "ACCEPTED": "Принята к публикации",
    "PUBLISHED": "Опубликована",
    "REJECTED": "Отклонена",
}


REPORT_VERIFICATION_STATUSES = {
    "PENDING": "Ожидает проверки",
    "APPROVED": "Проверена",
    "RETURNED": "Возвращена",
    "DUPLICATE": "Дубликат",
}


reports_bp = Blueprint(
    "reports",
    __name__,
    url_prefix="/reports",
)


def get_report_departments():
    query = (
        select(Department)
        .where(
            Department.is_active.is_(True),
        )
        .order_by(
            Department.name,
        )
    )

    if current_user.role == "DEPARTMENT_HEAD":
        if current_user.department_id is None:
            abort(403)

        query = query.where(
            Department.id == current_user.department_id,
        )

    return db.session.scalars(query).all()


def get_report_years():
    query = (
        select(Plan.year)
        .where(
            Plan.status == "APPROVED",
        )
        .distinct()
        .order_by(
            Plan.year.desc(),
        )
    )

    return db.session.scalars(query).all()


def get_report_reference_values(reference_type):
    return db.session.scalars(
        select(ReferenceValue)
        .where(
            ReferenceValue.type == reference_type,
            ReferenceValue.is_active.is_(True),
        )
        .order_by(
            ReferenceValue.sort_order,
            ReferenceValue.name_ru,
        )
    ).all()


def get_publications_report_filters():
    year_raw = request.args.get(
        "year",
        "",
    ).strip()

    department_raw = request.args.get(
        "department_id",
        "",
    ).strip()

    quartile_raw = request.args.get(
        "quartile_id",
        "",
    ).strip()

    indexing_type_raw = request.args.get(
        "indexing_type_id",
        "",
    ).strip()

    publication_status = request.args.get(
        "publication_status",
        "",
    ).strip()

    verification_status = request.args.get(
        "verification_status",
        "",
    ).strip()

    year = None
    department_id = None
    quartile_id = None
    indexing_type_id = None

    if year_raw:
        try:
            year = int(year_raw)
        except ValueError:
            abort(400)

    if department_raw:
        try:
            department_id = int(department_raw)
        except ValueError:
            abort(400)

    if quartile_raw:
        try:
            quartile_id = int(quartile_raw)
        except ValueError:
            abort(400)

    if indexing_type_raw:
        try:
            indexing_type_id = int(indexing_type_raw)
        except ValueError:
            abort(400)

    if publication_status:
        if publication_status not in REPORT_PUBLICATION_STATUSES:
            abort(400)

    if verification_status:
        if verification_status not in REPORT_VERIFICATION_STATUSES:
            abort(400)

    if (
        current_user.role == "DEPARTMENT_HEAD"
        and department_id is not None
        and department_id != current_user.department_id
    ):
        abort(403)

    return {
        "year": year,
        "department_id": department_id,
        "quartile_id": quartile_id,
        "indexing_type_id": indexing_type_id,
        "publication_status": (
            publication_status or None
        ),
        "verification_status": (
            verification_status or None
        ),
    }


def build_publications_report_query(
    year=None,
    department_id=None,
    publication_status=None,
    verification_status=None,
    quartile_id=None,
    indexing_type_id=None,
):
    query = (
        select(Publication)
        .join(
            Publication.plan_item,
        )
        .join(
            PlanItem.plan,
        )
        .join(
            Publication.department,
        )
    )

    if year is not None:
        query = query.where(
            Plan.year == year,
        )

    if department_id is not None:
        query = query.where(
            Publication.department_id == department_id,
        )

    if publication_status:
        query = query.where(
            Publication.status == publication_status,
        )

    if verification_status:
        query = query.where(
            Publication.verification_status
            == verification_status,
        )

    if quartile_id is not None:
        query = query.where(
            Publication.quartile_id == quartile_id,
        )

    if indexing_type_id is not None:
        query = query.where(
            Publication.indexing_type_id == indexing_type_id,
        )

    if current_user.role == "DEPARTMENT_HEAD":
        if current_user.department_id is None:
            abort(403)

        query = query.where(
            Publication.department_id
            == current_user.department_id,
        )

    return query.order_by(
        Plan.year.desc(),
        Publication.department_id,
        Publication.title,
        Publication.id,
    )


def get_quartile_name(publication):
    if publication.quartile is None:
        return ""

    if current_user.language == "kk":
        return publication.quartile.name_kk

    return publication.quartile.name_ru


def get_indexing_type_name(publication):
    if publication.indexing_type is None:
        return ""

    if current_user.language == "kk":
        return publication.indexing_type.name_kk

    return publication.indexing_type.name_ru


def get_publication_report_row(publication):
    plan = publication.plan_item.plan

    verifier = ""

    if publication.verifier:
        verifier = publication.verifier.full_name

    verified_at = ""

    if publication.verified_at:
        verified_at = publication.verified_at.strftime(
            "%d.%m.%Y %H:%M",
        )

    return [
        plan.year,
        publication.department.name,
        f"№{plan.id}",
        publication.title,
        publication.journal or "",
        publication.publication_year or "",
        publication.doi or "",
        get_quartile_name(publication),
        get_indexing_type_name(publication),
        REPORT_PUBLICATION_STATUSES.get(
            publication.status,
            publication.status,
        ),
        REPORT_VERIFICATION_STATUSES.get(
            publication.verification_status,
            publication.verification_status,
        ),
        verifier,
        verified_at,
    ]


def create_publications_workbook(publications):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Публикации"

    headers = [
        "Год плана",
        "Кафедра",
        "План",
        "Название публикации",
        "Журнал",
        "Год публикации",
        "DOI",
        "Квартиль",
        "Индексация",
        "Статус публикации",
        "Статус проверки",
        "Проверяющий",
        "Дата проверки",
    ]

    worksheet.append(headers)

    for publication in publications:
        worksheet.append(
            get_publication_report_row(
                publication,
            )
        )

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 2:
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(str(cell.value)),
            )

        column_letter = get_column_letter(
            column_cells[0].column,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 2, 12),
            60,
        )

    return workbook


def get_pdf_font_paths():
    regular_candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
    ]

    bold_candidates = [
        Path("C:/Windows/Fonts/arialbd.ttf"),
        Path("C:/Windows/Fonts/calibrib.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
    ]

    regular_path = next(
        (path for path in regular_candidates if path.is_file()),
        None,
    )
    bold_path = next(
        (path for path in bold_candidates if path.is_file()),
        None,
    )

    if regular_path is None or bold_path is None:
        raise RuntimeError(
            "Не найден Unicode-шрифт для генерации PDF. "
            "Установите Arial, Calibri или DejaVu Sans."
        )

    return regular_path, bold_path


def register_pdf_fonts():
    regular_path, bold_path = get_pdf_font_paths()

    pdfmetrics.registerFont(
        TTFont(
            "PublicationReportRegular",
            str(regular_path),
        )
    )

    pdfmetrics.registerFont(
        TTFont(
            "PublicationReportBold",
            str(bold_path),
        )
    )


def create_publications_pdf(publications):
    register_pdf_fonts()

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=7 * mm,
        leftMargin=7 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
        title="Отчёт по публикациям",
        author="Publication Monitor",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "PublicationReportTitle",
        parent=styles["Title"],
        fontName="PublicationReportBold",
        fontSize=14,
        leading=17,
        alignment=TA_LEFT,
        spaceAfter=5 * mm,
    )

    header_style = ParagraphStyle(
        "PublicationReportHeader",
        fontName="PublicationReportBold",
        fontSize=5.5,
        leading=6.5,
        alignment=TA_LEFT,
    )

    cell_style = ParagraphStyle(
        "PublicationReportCell",
        fontName="PublicationReportRegular",
        fontSize=5.2,
        leading=6.2,
        alignment=TA_LEFT,
    )

    story = [
        Paragraph(
            "Отчёт по публикациям",
            title_style,
        ),
        Spacer(
            1,
            2 * mm,
        ),
    ]

    headers = [
        "Год",
        "Кафедра",
        "План",
        "Публикация",
        "Журнал",
        "Год",
        "DOI",
        "Квартиль",
        "Индексация",
        "Статус",
        "Проверка",
        "Проверяющий",
        "Дата",
    ]

    rows = [
        [
            Paragraph(
                str(header),
                header_style,
            )
            for header in headers
        ]
    ]

    for publication in publications:
        values = get_publication_report_row(
            publication,
        )

        rows.append(
            [
                Paragraph(
                    str(value) if value != "" else "—",
                    cell_style,
                )
                for value in values
            ]
        )

    column_widths = [
        10 * mm,
        26 * mm,
        10 * mm,
        40 * mm,
        25 * mm,
        10 * mm,
        25 * mm,
        16 * mm,
        23 * mm,
        24 * mm,
        22 * mm,
        26 * mm,
        21 * mm,
    ]

    table = Table(
        rows,
        colWidths=column_widths,
        repeatRows=1,
        hAlign="LEFT",
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    "#E8E8E8",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.3,
                    "#999999",
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    2,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    2,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    2,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    2,
                ),
            ]
        )
    )

    story.append(table)

    story.append(
        Spacer(
            1,
            4 * mm,
        )
    )

    story.append(
        Paragraph(
            f"Всего публикаций: {len(publications)}",
            cell_style,
        )
    )

    document.build(story)

    output.seek(0)

    return output


@reports_bp.route("/publications")
@roles_required(
    "ADMIN",
    "DEPARTMENT_HEAD",
)
def publications_report():
    filters = get_publications_report_filters()

    query = build_publications_report_query(
        **filters,
    )

    publications = db.session.scalars(
        query,
    ).all()

    departments = get_report_departments()

    years = get_report_years()

    quartiles = get_report_reference_values(
        "QUARTILE",
    )

    indexing_types = get_report_reference_values(
        "INDEXING_TYPE",
    )

    return render_template(
        "reports/publications.html",
        publications=publications,
        departments=departments,
        years=years,
        quartiles=quartiles,
        indexing_types=indexing_types,
        publication_statuses=REPORT_PUBLICATION_STATUSES,
        verification_statuses=REPORT_VERIFICATION_STATUSES,
        selected_year=filters["year"],
        selected_department_id=filters[
            "department_id"
        ],
        selected_quartile_id=filters[
            "quartile_id"
        ],
        selected_indexing_type_id=filters[
            "indexing_type_id"
        ],
        selected_publication_status=filters[
            "publication_status"
        ],
        selected_verification_status=filters[
            "verification_status"
        ],
    )


@reports_bp.route("/publications/export.xlsx")
@roles_required(
    "ADMIN",
    "DEPARTMENT_HEAD",
)
def export_publications_excel():
    filters = get_publications_report_filters()

    query = build_publications_report_query(
        **filters,
    )

    publications = db.session.scalars(
        query,
    ).all()

    workbook = create_publications_workbook(
        publications,
    )

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="publication_report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

@reports_bp.route("/publications/export.pdf")
@roles_required(
    "ADMIN",
    "DEPARTMENT_HEAD",
)
def export_publications_pdf():
    filters = get_publications_report_filters()

    query = build_publications_report_query(
        **filters,
    )

    publications = db.session.scalars(
        query,
    ).all()

    output = create_publications_pdf(
        publications,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="publication_report.pdf",
        mimetype="application/pdf",
    )
